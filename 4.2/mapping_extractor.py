import json
import openpyxl
from openpyxl.styles import Border, Side, Alignment,  Font
from kuma_package import decrypt, password_to_key
import os
import argparse


def extract_enrichment(raw_enrichment):
    enrichment = []
    if len(raw_enrichment) != 0:
        for fields in raw_enrichment:
            enrichment.append({'sourceKind': fields['sourceKind'],
                               'targetField': fields['targetField'],
                               'sourceField': fields['sourceField'],
                               'constant': fields['constant'],
                               'template': fields['template'],
                               'sourceName': fields['sourceName'],
                               'keyFields': fields['keyFields'],
                               'mutation': [fields['mutation'][x]['kind'] for x in range(0, len(fields['mutation']))]
                               })
    sorted_enrichment = sorted(enrichment, key=lambda d: d['sourceKind'])
    return sorted_enrichment


def extract_mapping(raw_mapping):
    mapping = []
    if len(raw_mapping) != 0:
        for fields in raw_mapping:
            mapping.append({'sourceField': fields['sourceField'],
                            'eventField': fields['eventField'],
                            'label': fields['label']})
    return mapping


def prepare_normalizer(extras, normalizer):
    for extra in extras:
        normalizer.append({"name": extra['normalizer']['name'],
                           "mapping": extract_mapping(extra['normalizer']['mapping']),
                           "enrichment": extract_enrichment(extra['normalizer']['enrichment'])})
        if len(extra['normalizer']['extra']) != 0:
            prepare_normalizer(extra['normalizer']['extra'], normalizer)
    return normalizer


def sheet_style(active_sheet):
    for i in ['A', 'B', 'C']:
        column = active_sheet.column_dimensions[i]
        column.width = 40
        column.font = Font(name='Verdana', size=12)


def style_header(cell):
    cell.font = Font(name='Verdana', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center')


def style_table(active_sheet, start, end, start_letter, end_letter):
    cell_range = f'{start_letter}{start}:{end_letter}{end}'
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in active_sheet[cell_range]:
        for cell in row:
            cell.font = Font(name='Verdana', size=12, bold=False)
            cell.alignment = Alignment(wrap_text=True)
            cell.border = border


def style_table_header(active_sheet, start, start_letter, end_letter):
    header_range = f'{start_letter}{start}:{end_letter}{start}'

    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    for row in active_sheet[header_range]:
        for cell in row:
            cell.font = Font(name='Verdana', size=12, bold=True)
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
            cell.border = border


def insert_header(active_sheet, counter, header):
    current_cell = active_sheet.cell(row=counter, column=1)
    current_cell.value = header
    active_sheet.merge_cells(start_row=counter, start_column=1, end_row=counter, end_column=3)
    style_header(current_cell)
    return counter + 2


def insert_mapping(active_sheet, counter, mapping):
    start = counter
    active_sheet.cell(row=counter, column=1).value = 'Source Field'
    active_sheet.cell(row=counter, column=2).value = 'Target Field'
    active_sheet.cell(row=counter, column=3).value = 'Label'
    counter = counter + 1

    for maps in mapping:
        active_sheet.cell(row=counter, column=1).value = maps['sourceField']
        active_sheet.cell(row=counter, column=2).value = maps['eventField']
        active_sheet.cell(row=counter, column=3).value = maps['label']
        counter = counter + 1
    end = counter - 1
    style_table_header(active_sheet, start, 'A', 'C')
    style_table(active_sheet, start + 1, end, 'A', 'C')

    return counter + 1


def insert_enrichment_header(counter, kind, active_sheet):
    cell = active_sheet.cell(row=counter, column=1)
    cell.value = kind.title()
    cell.font = Font(name='Verdana', size=12, bold=True)
    counter = counter + 2
    end_letter = 'C'

    if kind == "event":
        active_sheet.cell(row=counter, column=1).value = 'Target Field'
        active_sheet.cell(row=counter, column=2).value = 'Source Field'
        active_sheet.cell(row=counter, column=3).value = 'Mutation'

    if kind == "template":
        active_sheet.cell(row=counter, column=1).value = 'Target Field'
        active_sheet.cell(row=counter, column=2).value = 'Template'
        end_letter = 'B'

    if kind == "constant":
        active_sheet.cell(row=counter, column=1).value = 'Target Field'
        active_sheet.cell(row=counter, column=2).value = 'Value'
        end_letter = 'B'

    if kind == "dictionary":
        active_sheet.cell(row=counter, column=1).value = 'Target Field'
        active_sheet.cell(row=counter, column=2).value = 'Key Fields'
        active_sheet.cell(row=counter, column=3).value = 'Dictionary'

    style_table_header(active_sheet, counter, 'A', end_letter)

    return counter + 1


def insert_enrichment_table(enrichment, counter, kind, active_sheet, dicts):
    end_letter = 'C'
    if kind == "event":
        active_sheet.cell(row=counter, column=1).value = enrichment['targetField']
        active_sheet.cell(row=counter, column=2).value = enrichment['sourceField']
        active_sheet.cell(row=counter, column=3).value = ','.join(set(enrichment['mutation']))

    if kind == "template":
        active_sheet.cell(row=counter, column=1).value = enrichment['targetField']
        active_sheet.cell(row=counter, column=2).value = enrichment['template']
        end_letter = 'B'

    if kind == "constant":
        active_sheet.cell(row=counter, column=1).value = enrichment['targetField']
        active_sheet.cell(row=counter, column=2).value = enrichment['constant']
        end_letter = 'B'

    if kind == "dictionary":
        active_sheet.cell(row=counter, column=1).value = enrichment['targetField']
        active_sheet.cell(row=counter, column=2).value = ','.join(enrichment['keyFields'])
        active_sheet.cell(row=counter, column=3).value = dicts[enrichment['sourceName']]

    style_table(active_sheet, counter, counter, 'A', end_letter)

    return counter + 1


def insert_enrichment(active_sheet, counter, enrichment, dicts):
    counter = counter + 1
    cell = active_sheet.cell(row=counter, column=2)
    cell.value = "Enrichments"
    style_header(cell)
    counter = counter + 1
    previous_kind = ""
    for fields in enrichment:

        kind = fields['sourceKind']

        if kind != previous_kind:
            counter = insert_enrichment_header(counter + 1, kind, active_sheet)

        counter = insert_enrichment_table(fields, counter, kind, active_sheet, dicts)
        previous_kind = kind

    return counter + 2


def insert_normalizer(normalizer, active_sheet, dicts):
    counter = 1
    for subparser in normalizer:
        counter = insert_header(active_sheet, counter, subparser['name'])
        counter = insert_mapping(active_sheet, counter, subparser['mapping'])
        if len(subparser['enrichment']) != 0:
            counter = insert_enrichment(active_sheet, counter, subparser['enrichment'], dicts)


if __name__ == '__main__':
    # args
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', help='file with resources', required=True)
    parser.add_argument('-p', help='password for resources', required=True)
    args = parser.parse_args()

    input_file = args.f
    password = args.p

    # prepare normalizer
    decrypt(input_file, 'resources.json.temp', password_to_key(password), pretty=False)

    with open('resources.json.temp', 'r') as normalizerFile:
        raw_resources = json.load(normalizerFile)

    # for id to name mapping
    dictionaries = {}
    for resource in raw_resources['resources']:
        if resource['kind'] == 'dictionary':
            dictionaries.update({resource['id']: resource['name']})

    for resource in raw_resources['resources']:
        if resource['kind'] == 'normalizer':
            main_normalizer = [{"name": resource['encoded']['payload']['name'],
                                "mapping": extract_mapping(resource['encoded']['payload']['mapping']),
                                "enrichment": extract_enrichment(resource['encoded']['payload']['enrichment'])}]
            prepared_normalizer = prepare_normalizer(resource['encoded']['payload']['extra'], main_normalizer)

            # work with Excel
            wb = openpyxl.Workbook()

            sheet = wb.active
            sheet_style(sheet)
            insert_normalizer(prepared_normalizer, sheet, dictionaries)

            forbidden_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
            name_to_save = resource['name']
            for ch in forbidden_chars:
                name_to_save = name_to_save.replace(ch, ' ')

            wb.save('_' + name_to_save + '.xlsx')

    os.remove('resources.json.temp')
